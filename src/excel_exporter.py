import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import io

class ExcelExporter:
    """Classe pour exporter les données vers Excel avec mise en forme"""
    
    def __init__(self, df: pd.DataFrame):
        self.df = df
    
    def export_basic(self, output_path: str) -> str:
        """Export Excel simple"""
        self.df.to_excel(output_path, index=False, sheet_name="Données")
        return output_path
    
    def export_formatted(self, output_path: str) -> str:
        """Export Excel avec mise en forme professionnelle"""
        
        # Création du fichier
        self.df.to_excel(output_path, index=False, sheet_name="Données")
        
        # Chargement pour mise en forme
        wb = load_workbook(output_path)
        ws = wb.active
        
        # Style header
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Appliquer le style aux en-têtes
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Ajuster largeur des colonnes
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze first row
        ws.freeze_panes = "A2"
        
        wb.save(output_path)
        return output_path
    
    def get_download_buffer(self) -> io.BytesIO:
        """Retourne un buffer pour téléchargement direct dans Streamlit"""
        buffer = io.BytesIO()
        self.df.to_excel(buffer, index=False, sheet_name="Données")
        buffer.seek(0)
        return buffer
